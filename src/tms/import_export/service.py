from __future__ import annotations

from collections.abc import Iterable, Iterator
from pathlib import Path
from typing import Any

from ..core.enums import JobState, JobType
from ..core.events import DomainEvent
from ..datasets.models import Dataset
from ..datasets.repository import DatasetRepository
from ..jobs.models import Job
from ..jobs.repository import JobRepository
from ..members.models import Member
from ..members.repository import MemberRepository
from ..runtime.event_bus import EventBus
from .csv_exporter import export_csv
from .csv_importer import iter_csv_members
from .xlsx_exporter import export_xlsx
from .xlsx_importer import iter_xlsx_members


class ImportExportService:
    def __init__(
        self,
        datasets: DatasetRepository,
        members: MemberRepository,
        jobs: JobRepository,
        events: EventBus,
    ) -> None:
        self.datasets = datasets
        self.members = members
        self.jobs = jobs
        self.events = events

    @staticmethod
    def _member_iterator(path: Path) -> Iterator[Member]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return iter_csv_members(path)
        if suffix in {".xlsx", ".xlsm"}:
            return iter_xlsx_members(path)
        raise ValueError("Only CSV and XLSX files are supported")

    def import_dataset(
        self,
        path: Path,
        name: str,
        *,
        account_id: int | None = None,
        batch_size: int = 1000,
    ) -> tuple[int, int]:
        if not path.exists():
            raise FileNotFoundError(path)
        dataset_id = self.datasets.create(
            Dataset(
                id=None,
                name=name or path.stem,
                source_type=path.suffix.lower().lstrip(".").upper(),
                source_reference=str(path),
            )
        )
        job_id = self.jobs.create(
            Job(
                id=None,
                job_type=JobType.IMPORT,
                state=JobState.RUNNING,
                account_id=account_id,
                source_dataset_id=dataset_id,
            )
        )
        accepted = 0
        invalid = 0
        batch: list[Member] = []
        self.jobs.submit_event(
            job_id,
            "INFO",
            "IMPORT_STARTED",
            f"path={path} dataset_id={dataset_id}",
            critical=True,
        ).result(timeout=10.0)
        try:
            for member in self._member_iterator(path):
                batch.append(member)
                if len(batch) >= batch_size:
                    summary = self.members.submit_ingest_batch(
                        dataset_id,
                        batch,
                        account_id=account_id,
                        source_label=path.name,
                    ).result(timeout=60.0)
                    accepted += summary.accepted
                    invalid += summary.invalid
                    batch = []
                    self.events.publish(
                        DomainEvent(
                            "ImportProgress",
                            {
                                "job_id": job_id,
                                "dataset_id": dataset_id,
                                "accepted": accepted,
                                "invalid": invalid,
                            },
                        )
                    )
            if batch:
                summary = self.members.submit_ingest_batch(
                    dataset_id,
                    batch,
                    account_id=account_id,
                    source_label=path.name,
                ).result(timeout=60.0)
                accepted += summary.accepted
                invalid += summary.invalid
            self.jobs.submit_set_state(
                job_id,
                JobState.COMPLETED_WITH_ERRORS if invalid else JobState.COMPLETED,
                checkpoint={"accepted": accepted, "invalid": invalid},
            ).result(timeout=10.0)
            self.jobs.submit_event(
                job_id,
                "INFO",
                "IMPORT_COMPLETED",
                f"accepted={accepted} invalid={invalid}",
                critical=True,
            ).result(timeout=10.0)
            self.events.publish(
                DomainEvent(
                    "ImportCompleted",
                    {
                        "job_id": job_id,
                        "dataset_id": dataset_id,
                        "accepted": accepted,
                        "invalid": invalid,
                    },
                )
            )
            return dataset_id, job_id
        except Exception as exc:
            self.jobs.submit_set_state(
                job_id,
                JobState.FAILED,
                checkpoint={"accepted": accepted, "invalid": invalid, "error": str(exc)},
            ).result(timeout=10.0)
            self.jobs.submit_event(
                job_id,
                "ERROR",
                "IMPORT_FAILED",
                str(exc),
                critical=True,
            ).result(timeout=10.0)
            raise

    @staticmethod
    def _flatten_chunks(chunks: Iterable[list[dict]]) -> Iterator[dict]:
        for chunk in chunks:
            yield from chunk

    def export_dataset(
        self,
        dataset_id: int,
        path: Path,
        *,
        account_id: int | None = None,
    ) -> int:
        job_id = self.jobs.create(
            Job(
                id=None,
                job_type=JobType.EXPORT,
                state=JobState.RUNNING,
                account_id=account_id,
                source_dataset_id=dataset_id,
            )
        )
        self.jobs.submit_event(
            job_id,
            "INFO",
            "EXPORT_STARTED",
            f"dataset_id={dataset_id} path={path}",
            critical=True,
        ).result(timeout=10.0)
        try:
            rows = self._flatten_chunks(
                self.datasets.iter_export_rows(dataset_id, account_id=account_id)
            )
            if path.suffix.lower() == ".csv":
                export_csv(path, rows)
            elif path.suffix.lower() == ".xlsx":
                export_xlsx(path, rows)
            else:
                raise ValueError("Export path must end with .csv or .xlsx")
            self.jobs.submit_set_state(job_id, JobState.COMPLETED).result(timeout=10.0)
            self.jobs.submit_event(
                job_id,
                "INFO",
                "EXPORT_COMPLETED",
                f"path={path}",
                critical=True,
            ).result(timeout=10.0)
            self.events.publish(
                DomainEvent(
                    "ExportCompleted",
                    {"job_id": job_id, "dataset_id": dataset_id, "path": str(path)},
                )
            )
            return job_id
        except Exception as exc:
            self.jobs.submit_set_state(job_id, JobState.FAILED).result(timeout=10.0)
            self.jobs.submit_event(
                job_id,
                "ERROR",
                "EXPORT_FAILED",
                str(exc),
                critical=True,
            ).result(timeout=10.0)
            raise

    def export_job_log(self, job_id: int, path: Path) -> None:
        rows = self.jobs.event_rows(job_id, limit=10000)
        path.parent.mkdir(parents=True, exist_ok=True)
        fields = ["id", "timestamp", "level", "event_code", "member_id", "message"]
        if path.suffix.lower() == ".csv":
            import csv

            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fields)
                writer.writeheader()
                writer.writerows(rows)
            return
        if path.suffix.lower() == ".xlsx":
            from openpyxl import Workbook

            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet("Job Log")
            sheet.append(fields)
            for row in rows:
                sheet.append([row.get(field) for field in fields])
            workbook.save(path)
            return
        raise ValueError("Log export path must end with .csv or .xlsx")

    def export_job_results(self, job_id: int, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        fields = [
            "ordinal",
            "target_state",
            "state",
            "attempt_count",
            "last_error_code",
            "last_error_text",
            "processed_at",
            "telegram_user_id",
            "username",
            "first_name",
            "last_name",
            "phone",
        ]
        chunks = self.jobs.iter_export_result_rows(job_id)
        if path.suffix.lower() == ".csv":
            import csv

            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fields)
                writer.writeheader()
                for chunk in chunks:
                    writer.writerows(chunk)
            return
        if path.suffix.lower() == ".xlsx":
            from openpyxl import Workbook

            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet("Kết quả xử lý")
            sheet.append(fields)
            for chunk in chunks:
                for row in chunk:
                    sheet.append([row.get(field) for field in fields])
            workbook.save(path)
            return
        raise ValueError("Result export path must end with .csv or .xlsx")

