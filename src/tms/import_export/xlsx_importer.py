from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from ..members.models import Member
from .column_mapper import map_headers
from .parsing import member_from_mapping


def iter_xlsx_members(path: Path) -> Iterator[Member]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "") for value in next(rows)]
        except StopIteration:
            return
        mapping = map_headers(headers)
        if not mapping:
            raise ValueError("XLSX has no recognized member columns")
        for row in rows:
            data = {
                field: (row[index] if index < len(row) else None)
                for index, field in mapping.items()
            }
            yield member_from_mapping(data)
    finally:
        workbook.close()


def import_xlsx(path: Path) -> list[Member]:
    return list(iter_xlsx_members(path))
